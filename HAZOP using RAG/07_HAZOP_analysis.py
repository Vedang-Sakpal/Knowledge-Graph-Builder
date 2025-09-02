import json
import re
from typing import Dict, List, Any, Optional
from datetime import datetime

from neo4j import GraphDatabase
import google.generativeai as genai
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Assuming config.py exists with your API keys and Neo4j credentials
import config


def get_driver() -> GraphDatabase.driver:
    """Establishes connection to the Neo4j database."""
    return GraphDatabase.driver(config.NEO4J_URI, auth=(config.NEO4J_USERNAME, config.NEO4J_PASSWORD))


def fetch_equipment_nodes(session) -> List[Dict[str, str]]:
    """Fetches all equipment nodes from the Neo4j database."""
    query = """
    MATCH (e:Equipment)
    RETURN e.name AS name, e.type AS type
    ORDER BY name
    """
    result = session.run(query)
    return [dict(record) for record in result]


def fetch_equipment_context(session, equipment_name: str) -> dict:
    """
    Fetches the context for a specific piece of equipment, including connected
    sensors and valves.
    """
    q = """
    MATCH (e:Equipment {name: $name})
    OPTIONAL MATCH (s:Sensor)-[:CONNECTED_TO]->(e)
    OPTIONAL MATCH (v_in:Valve)-[:FLOWS_TO]->(e)
    OPTIONAL MATCH (e)-[:FLOWS_TO]->(v_out:Valve)
    WITH e, 
         collect(DISTINCT s {name: s.name, type: s.type}) AS sensors, 
         collect(DISTINCT v_in {name: v_in.name, type: v_in.type}) AS incoming_valves,
         collect(DISTINCT v_out {name: v_out.name, type: v_out.type}) AS outgoing_valves
    RETURN {
        equipment: e.name,
        type: e.type,
        sensors: sensors,
        incoming_valves: incoming_valves,
        outgoing_valves: outgoing_valves
    } AS context
    """
    result = session.run(q, name=equipment_name).single()
    return result['context'] if result else {"sensors": [], "valves": []}


def load_applicable_deviations(equipment_name: str) -> List[Dict[str, str]]:
    """Load applicable deviations for the selected equipment from the CSV."""
    deviations = []
    try:
        # NOTE: This still reads from a CSV, as per the original design.
        with open(config.APPLICABLE_DEVIATIONS_PATH, newline="", encoding="utf-8") as f:
            # Using standard library's csv reader for simplicity here
            import csv
            reader = csv.DictReader(f)
            for row in reader:
                if row["EquipmentName"] == equipment_name:
                    deviations.append({
                        "parameter": row["Parameter"],
                        "guideword": row["Guideword"],
                        "rationale": row["Rationale"]
                    })
    except FileNotFoundError:
        print("Warning: Applicable_Deviations.csv not found. Will use generic deviations.")
    
    return deviations


def load_process_description() -> str:
    """Load the generated process description."""
    try:
        with open(config.GENERATED_PROCESS_DESCRIPTION_PATH, "r", encoding="utf-8") as f:
            return f.read()
    except FileNotFoundError:
        return "Process description not available."


def configure_llm():
    """Configures and returns the Generative AI model."""
    genai.configure(api_key=config.GEMINI_API_KEY)
    return genai.GenerativeModel(config.GEMINI_MODEL_NAME)


def save_and_parse_llm_json(llm_raw_text: str, equipment_name: str) -> Optional[List[Dict[str, Any]]]:
    """
    Saves the raw LLM text output to a file, then cleans and parses it into a structured list.
    This prevents parsing errors by isolating the raw output first.
    """
    # Define the filename for the raw output
    raw_filename = f"Results/HAZOP_Analysis_{equipment_name.replace(' ', '_')}_raw_output.json"

    # --- Step 1: Clean the raw text ---
    # Remove markdown code fences (e.g., ```json ... ```) and strip whitespace
    cleaned_text = re.sub(r'```json\s*|```', '', llm_raw_text).strip()

    # --- Step 2: Save the cleaned, raw JSON output to a file ---
    try:
        with open(raw_filename, "w", encoding="utf-8") as f:
            f.write(cleaned_text)
        print(f"   -> LLM raw output saved to: {raw_filename}")
    except IOError as e:
        print(f"Error saving raw LLM output file: {e}")
        return None

    # --- Step 3: Parse the cleaned text into a Python object ---
    try:
        data = json.loads(cleaned_text)
        
        # --- Step 4: Normalize the data structure to ensure consistency ---
        if not isinstance(data, list):
            print("Warning: LLM output was not a list as expected. The output will be wrapped in a list.")
            data = [data]

        normalized = []
        for item in data:
            if not isinstance(item, dict):
                continue  # Skip any non-dictionary items in the list
            normalized.append({
                "parameter": item.get("parameter", "N/A"),
                "guideword": item.get("guideword", "N/A"),
                "causes": item.get("causes") if isinstance(item.get("causes"), list) else [str(item.get("causes", "N/A"))],
                "consequences": item.get("consequences") if isinstance(item.get("consequences"), list) else [str(item.get("consequences", "N/A"))],
                "safeguards": item.get("safeguards") if isinstance(item.get("safeguards"), list) else [str(item.get("safeguards", "N/A"))],
                "recommendations": item.get("recommendations") if isinstance(item.get("recommendations"), list) else [str(item.get("recommendations", "N/A"))]
            })
        return normalized
    except json.JSONDecodeError as e:
        print(f"❌ Critical Error: Failed to parse JSON from the LLM's response. {e}")
        print(f"   Please check the raw output file for issues: {raw_filename}")
        return None


def conduct_hazop_analysis(model, equipment: Dict[str, str], context: Dict[str, Any], 
                          deviations: List[Dict[str, str]], process_description: str) -> Optional[List[Dict[str, Any]]]:
    """
    Conducts HAZOP analysis using the LLM and returns the parsed, structured result.
    """

    prompt = {
            "analysis_task": """
            Perform a detailed Hazard and Operability (HAZOP) analysis for the specified equipment using the provided P&ID context and process description. Analyze the given deviations and generate the results strictly following the rules below.
        
            ### Rules for Identifying Causes:
            1.  **Instrument & Control Failure:** Analyze causes arising from the failure of any specified instruments, control loops, and alarms associated with the equipment.
            2.  **Process & Material Issues:** Based on operating conditions, design parameters, and chemical safety data, identify causes related to equipment damage , blockages, or leaks. Consider material compatibility and chemical reactions.
            3.  **Stream Interactions:** For equipment with multiple streams, consider how interactions between streams could cause deviations.
            4.  **Human Factors:** Include potential causes from human error, such as incorrect operator actions or maintenance mistakes.
            5.  **Strict Constraint:** Do not assume the existence of any equipment or instrument that is not explicitly described in the provided context.
        
            ### Rules for Determining Consequences:
            1.  **Chemical Hazards:** Evaluate consequences based on the properties of the chemicals involved (e.g., toxicity, flammability).
            2.  **Equipment Damage:** Describe any potential damage to the equipment itself resulting from the deviation.
            3.  **HSE Impact:** Detail the consequences for personnel safety, human health, and the environment (HSE).
        
            ### Rules for Safeguards & Recommendations:
            1.  **Existing Safeguards:** In the 'safeguards' list, only include protective devices, alarms, and procedures that are explicitly present in the provided system information.
            2.  **Recommendations:** In the 'recommendations' list, identify and suggest necessary safeguards that are missing from the current system but would mitigate the identified risk.
            """,
            "equipment": {
                "name": equipment.get("name"),
                "type": equipment.get("type"),
            },
            "p&id_context": context,
            "process_description": process_description,
            "deviations_to_analyze": deviations,
            "required_output_format": "[{\"parameter\": \"...\", \"guideword\": \"...\", \"causes\": [\"...\"], \"consequences\": [\"...\"], \"safeguards\": [\"...\"], \"recommendations\": [\"...\"]}, ...]"
        }

    response = model.generate_content(
        [
            {"role": "user", "parts": [json.dumps(prompt, indent=2)]}
        ]
    )
    raw_text = (response.text or "").strip()
    
    parsed_analysis = save_and_parse_llm_json(raw_text, equipment['name'])
    
    return parsed_analysis
        

def display_equipment_list(equipments: List[Dict[str, str]]):
    """Display numbered list of available equipment."""
    print("\n=== Available Equipment ===")
    for i, eq in enumerate(equipments, 1):
        print(f"{i}. {eq['name']} ({eq['type']})")


def get_equipment_selection(equipments: List[Dict[str, str]]) -> Optional[Dict[str, str]]:
    """Get equipment selection from user."""
    while True:
        try:
            choice = input("\nEnter equipment number (or 'q' to quit): ").strip()
            if choice.lower() == 'q':
                return None
            
            choice_num = int(choice)
            if 1 <= choice_num <= len(equipments):
                return equipments[choice_num - 1]
            else:
                print(f"Please enter a number between 1 and {len(equipments)}")
        except ValueError:
            print("Please enter a valid number")


def save_hazop_report_excel(equipment_name: str, analysis_data: List[Dict[str, Any]]):
    """Save HAZOP analysis to a styled Excel file."""
    filename = f"Results/HAZOP_Analysis_{equipment_name.replace(' ', '_')}.xlsx"
    
    wb = Workbook()
    ws = wb.active
    ws.title = "HAZOP Analysis"

    if not analysis_data:
        print("⚠️ No valid analysis data provided. Excel report will contain an error message.")
        ws.append(["Error", "Failed to generate HAZOP analysis due to a parsing error. See console for details."])
        wb.save(filename)
        return

    # --- Define Styles ---
    header_font = Font(bold=True, color="FFFFFF", name="Calibri")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    
    thin_border = Border(
        left=Side(style='thin'), 
        right=Side(style='thin'), 
        top=Side(style='thin'), 
        bottom=Side(style='thin')
    )
    
    # Alignment for wrapping text in data cells
    wrap_alignment = Alignment(wrap_text=True, vertical='top', horizontal='left')

    # --- Write Header ---
    header = [
        "Equipment", "Parameter", "Guideword", "Deviation",
        "Causes", "Consequences", "Safeguards", "Recommendations", "Analysis Date"
    ]
    ws.append(header)
    
    # Apply styles to header row
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # --- Write Data Rows ---
    for d in analysis_data:
        parameter = d.get("parameter", "N/A")
        guideword = d.get("guideword", "N/A")
        deviation = f"{parameter} {guideword}".strip()
        
        # Join list items with a newline for better Excel readability
        causes = "\n".join(f"- {c}" for c in d.get("causes", []))
        consequences = "\n".join(f"- {c}" for c in d.get("consequences", []))
        safeguards = "\n".join(f"- {s}" for s in d.get("safeguards", []))
        recommendations = "\n".join(f"- {r}" for r in d.get("recommendations", []))
        
        row_data = [
            equipment_name, parameter, guideword, deviation,
            causes, consequences, safeguards, recommendations,
            datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        ]
        ws.append(row_data)

        # Apply border and wrap-text alignment to the new row
        for col_idx in range(1, len(row_data) + 1):
            cell = ws.cell(row=ws.max_row, column=col_idx)
            cell.border = thin_border
            # Apply wrap text to columns that need it
            if col_idx in [5, 6, 7, 8]: # Causes, Consequences, Safeguards, Recommendations
                cell.alignment = wrap_alignment

    # --- Adjust Column Widths ---
    column_widths = {'A': 20, 'B': 15, 'C': 15, 'D': 25, 'E': 50, 'F': 50, 'G': 50, 'H': 50, 'I': 20}
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    try:
        wb.save(filename)
        print(f"\n✅ HAZOP analysis saved to Excel: {filename}")
    except IOError as e:
        print(f"❌ Error saving Excel file: {e}")


def main():
    """Main function to run the HAZOP analysis tool."""
    print("=== Professional HAZOP Analysis Tool ===\n")
    
    model = configure_llm()
    process_description = load_process_description()
    
    driver = get_driver()
    try:
        with driver.session() as session:
            equipments = fetch_equipment_nodes(session)
            if not equipments:
                print("No equipment found in the database.")
                return
            
            while True:
                display_equipment_list(equipments)
                selected_equipment = get_equipment_selection(equipments)
                if selected_equipment is None:
                    break
                
                equipment_name = selected_equipment["name"]
                print(f"\n--- Starting HAZOP for: {equipment_name} ---\n")
                
                print("1. Fetching P&ID context from Neo4j...")
                context = fetch_equipment_context(session, equipment_name)
                
                print("2. Loading applicable deviations from CSV...")
                deviations = load_applicable_deviations(equipment_name)
                if not deviations:
                    print("   -> No specific deviations found. Using a generic deviation for analysis.")
                    deviations = [{"parameter": "Process", "guideword": "General Malfunction", "rationale": "Generic analysis due to lack of specific deviations."}]
                
                print("3. Conducting HAZOP analysis with LLM... (This may take a moment)")
                analysis = conduct_hazop_analysis(model, selected_equipment, context, deviations, process_description)

                if analysis:
                    print("4. Saving HAZOP report to Excel...")
                    save_hazop_report_excel(equipment_name, analysis)
                else:
                    print("\nSkipping report generation due to an error in the analysis step.")

                another = input("\nAnalyze another piece of equipment? (y/n): ").strip().lower()
                if another != 'y':
                    break
                
    finally:
        driver.close()
    
    print("\nHAZOP analysis complete!")


if __name__ == "__main__":
    main()